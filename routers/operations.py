from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import text
from typing import List, Optional
from urllib.parse import urlparse
from datetime import datetime, time, timedelta
import os
import io
import csv
import re
from pydantic import BaseModel
import pdfplumber
from docx import Document as DocxDocument
import openpyxl


from DB.database import get_db, MINIO_BUCKET_NAME
from DB.minio_client import get_minio_client
from DB.models.oms import (
    Operation as OperationModel,
    OperationDocument as OperationDocumentModel,
    ToolWithPart as ToolWithPartModel,
    PartType as PartTypeModel
)
from DB.models.configuration import workcenter as workcenterModel, Machine as MachineModel
from DB.models.inventory import Vendors
from DB.models.access_control import AccessUser
from DB.schemas.oms import Operation, OperationCreate, OperationUpdate
from services.notification_service import NotificationService

router = APIRouter(
    prefix="/operations",
    tags=["operations"]
)


class OperationPreview(BaseModel):
    operation_number: str
    operation_name: str
    setup_time: Optional[str] = None
    cycle_time: Optional[str] = None
    work_instructions: Optional[str] = None
    notes: Optional[str] = None


def _normalize_header(name: str) -> str:
    return name.strip().lower().replace("\n", " ").replace("\r", " ")


def _assert_part_not_schedule_active(part_id: int, db: Session) -> None:
    """Check if part has active schedule status and raise error if so."""
    schedule_status = db.execute(
        text("SELECT status FROM scheduling.part_schedule_status WHERE part_id = :pid"),
        {"pid": part_id}
    ).fetchone()
    if schedule_status and schedule_status[0] and schedule_status[0].lower() == "active":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Sorry, operations cannot be swapped because this part is currently scheduled for production. To swap operations, please inactivate the part's schedule status first."
        )


def _match_column(header: str) -> Optional[str]:
    h = _normalize_header(header)
    if "op" in h and ("number" in h or "num" in h or "#" in h):
        return "operation_number"
    if "operation" in h and "name" in h:
        return "operation_name"
    if "setup" in h:
        return "setup_time"
    if "cycle" in h:
        return "cycle_time"
    if "instruction" in h or "work inst" in h:
        return "work_instructions"
    if "note" in h:
        return "notes"
    return None


def _hms_from_seconds(total_seconds: int) -> str:
    sign = "-" if total_seconds < 0 else ""
    total_seconds = abs(int(total_seconds))
    hours, rem = divmod(total_seconds, 3600)
    minutes, seconds = divmod(rem, 60)
    return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"


_DAY_HMS_RE = re.compile(
    r"^(?:(\d+)\s+days?,\s*)?(-?\d+):(\d{1,2})(?::(\d{1,2})(?:\.\d+)?)?$",
    re.IGNORECASE,
)


def _format_duration_cell(value) -> Optional[str]:
    """Keep durations such as 100:00:00 instead of Excel's '4 days, 4:00:00'."""
    if value is None:
        return None
    if isinstance(value, timedelta):
        return _hms_from_seconds(int(value.total_seconds()))
    if isinstance(value, time):
        return f"{value.hour:02d}:{value.minute:02d}:{value.second:02d}"
    if isinstance(value, datetime):
        t = value.time()
        return f"{t.hour:02d}:{t.minute:02d}:{t.second:02d}"
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        days = float(value)
        # Excel stores time/duration as a fraction of a day (100h ≈ 4.1667)
        if 0 <= days < 5:
            return _hms_from_seconds(int(round(days * 24 * 3600)))
        text = str(value).strip()
    else:
        text = _sanitize_singleline_text(value)
    if not text:
        return None
    match = _DAY_HMS_RE.match(text)
    if match:
        days = int(match.group(1) or 0)
        hours = int(match.group(2))
        minutes = int(match.group(3))
        seconds = int(match.group(4) or 0)
        return _hms_from_seconds(days * 86400 + hours * 3600 + minutes * 60 + seconds)
    return text


def _is_zero_hms(val) -> bool:
    if val is None:
        return True
    if isinstance(val, time):
        return val.hour == 0 and val.minute == 0 and val.second == 0
    text = str(val).strip()
    if not text:
        return True
    parts = text.split(":")
    try:
        hours = int(parts[0])
        minutes = int(parts[1]) if len(parts) > 1 else 0
        seconds = int(str(parts[2]).split(".")[0]) if len(parts) > 2 else 0
        return hours == 0 and minutes == 0 and seconds == 0
    except (ValueError, IndexError):
        return False


def _sanitize_singleline_text(value) -> str:
    """Flatten cell text for op number, times, names, etc."""
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\(cid:\d+\)", " ", text)
    text = re.sub(r"_x000d_", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"_x000a_", " ", text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text.replace("\r", " ").replace("\n", " ")).strip()


def _sanitize_multiline_text(value) -> Optional[str]:
    """
    Preserve line breaks for work instructions / notes.
    Handles PDF (cid:13)/(cid:10), Excel _x000d_/_x000a_, and normal CR/LF.
    """
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None

    text = re.sub(r"\(cid:10\)", "\n", text)
    text = re.sub(r"\(cid:13\)", "\n", text)
    text = re.sub(r"_x000d_", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"_x000a_", "\n", text, flags=re.IGNORECASE)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # Collapse blank lines left by double PDF/Excel line-break markers
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip() or None


def _find_header_row(rows: list) -> int:
    """
    Scan rows top-to-bottom and return the index of the first row
    that contains BOTH an op-number column and an operation-name column.
    Returns -1 if not found.
    This lets us skip title rows, meta-info rows, blank rows etc.
    """
    for idx, row in enumerate(rows):
        mapped = set()
        for cell in row:
            key = _match_column(str(cell) if cell is not None else "")
            if key:
                mapped.add(key)
        if "operation_number" in mapped and "operation_name" in mapped:
            return idx
    return -1


def _parse_rows(rows: list) -> List[OperationPreview]:
    if not rows:
        return []

    # Find the actual header row (skip title / meta rows at top)
    header_idx = _find_header_row(rows)
    if header_idx == -1:
        return []

    header = rows[header_idx]
    header_map: dict[str, int] = {}
    for col_idx, cell in enumerate(header):
        key = _match_column(str(cell) if cell is not None else "")
        if key:
            header_map[key] = col_idx

    if "operation_number" not in header_map or "operation_name" not in header_map:
        return []

    result: List[OperationPreview] = []
    for row in rows[header_idx + 1:]:
        orig = list(row)
        while len(orig) <= max(header_map.values()):
            orig.append(None)
        raw_cells = [str(c) if c is not None else "" for c in orig]

        cells = [_sanitize_singleline_text(c) for c in raw_cells]

        # Skip completely empty rows and meta/note rows that slip through
        if not "".join(cells).strip():
            continue

        op_num  = cells[header_map["operation_number"]]
        op_name = cells[header_map["operation_name"]]

        # Skip rows where op_number or op_name look like note/footer lines
        if not op_num or not op_name:
            continue

        data = {
            "operation_number":  op_num,
            "operation_name":    op_name,
            "setup_time":        _format_duration_cell(orig[header_map["setup_time"]]) if "setup_time" in header_map else None,
            "cycle_time":        _format_duration_cell(orig[header_map["cycle_time"]]) if "cycle_time" in header_map else None,
            "work_instructions": _sanitize_multiline_text(raw_cells[header_map["work_instructions"]]) if "work_instructions" in header_map else None,
            "notes":             _sanitize_multiline_text(raw_cells[header_map["notes"]])             if "notes"             in header_map else None,
        }
        # Replace empty strings with None for optional fields
        for k in ("setup_time", "cycle_time", "work_instructions", "notes"):
            if data[k] == "":
                data[k] = None

        result.append(OperationPreview(**data))

    return result


# ── CSV ───────────────────────────────────────────────────────────────────────
def _parse_csv(content: bytes) -> List[OperationPreview]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    # Keep every row (including blanks) so _find_header_row can locate the header
    rows = list(reader)
    return _parse_rows(rows)


# ── XLSX ──────────────────────────────────────────────────────────────────────
def _parse_xlsx(content: bytes) -> List[OperationPreview]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        parsed = _parse_rows(rows)
        if parsed:
            return parsed
    return []


# ── DOCX ──────────────────────────────────────────────────────────────────────
def _parse_docx(content: bytes) -> List[OperationPreview]:
    doc = DocxDocument(io.BytesIO(content))
    for table in doc.tables:
        rows = [[cell.text for cell in row.cells] for row in table.rows]
        parsed = _parse_rows(rows)
        if parsed:
            return parsed
    return []


# ── PDF ───────────────────────────────────────────────────────────────────────
def _parse_pdf(content: bytes) -> List[OperationPreview]:
    operations: List[OperationPreview] = []
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                parsed = _parse_rows(table)
                if parsed:
                    operations.extend(parsed)
            if operations:
                break
    return operations


@router.post("/", response_model=Operation, status_code=status.HTTP_201_CREATED)
def create_operation(operation: OperationCreate, db: Session = Depends(get_db)):
    """Create a new operation"""
    data = operation.model_dump()
    part_type_id = data.get("part_type_id") or 1
    data["part_type_id"] = part_type_id

    # Check if part is scheduled (status is "active")
    part_id = data.get("part_id")
    if part_id:
        schedule_status = db.execute(
            text("SELECT status FROM scheduling.part_schedule_status WHERE part_id = :pid"),
            {"pid": part_id}
        ).fetchone()
        
        if schedule_status and schedule_status[0] and schedule_status[0].lower() == "active":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Sorry, this operation cannot be added/modified/deleted because the part is currently scheduled for production. To make changes, please inactivate the part's schedule status first."
            )

    # Validate required times only for non Out-Source operations
    setup_time_val = data.get("setup_time")
    cycle_time_val = data.get("cycle_time")
    if part_type_id != 2:
        if _is_zero_hms(setup_time_val) or _is_zero_hms(cycle_time_val):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="setup_time and cycle_time are mandatory and cannot be 00:00:00 for non Out-Source operations",
            )

    # Ensure part_id is present
    if not part_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="part_id is required for operations",
        )

    # Handle operation_number: ensure uniqueness per part and auto-generate when missing
    op_number_raw = data.get("operation_number")
    op_number = op_number_raw.strip() if isinstance(op_number_raw, str) else None

    if op_number:
        existing = (
            db.query(OperationModel)
            .filter(
                OperationModel.part_id == part_id,
                OperationModel.operation_number == op_number,
            )
            .first()
        )
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Operation number '{op_number}' already exists for this part",
            )
        data["operation_number"] = op_number
    else:
        existing_ops = (
            db.query(OperationModel)
            .filter(OperationModel.part_id == part_id)
            .all()
        )
        max_num = 0
        for op in existing_ops:
            try:
                n = int(str(op.operation_number).strip())
            except (TypeError, ValueError):
                continue
            if n > max_num:
                max_num = n
        next_num = max_num + 10 if max_num > 0 else 10
        data["operation_number"] = str(next_num)

    if part_type_id == 2:
        if not data.get("from_date") or not data.get("to_date"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Outsource operations require from_date and to_date",
            )
    db_operation = OperationModel(**data)
    db.add(db_operation)
    db.commit()
    db.refresh(db_operation)

    # Log operation creation for PC notifications
    user_name = None
    user_role = None
    if db_operation.user_id:
        user = db.query(AccessUser).filter(AccessUser.id == db_operation.user_id).first()
        user_name = user.user_name if user else None
        user_role = user.role if user else None
    
    NotificationService.log_operation_change(
        db=db,
        operation_id=db_operation.id,
        action="created",
        user_id=db_operation.user_id,
        user_name=user_name,
        user_role=user_role,
        details={"operation_name": db_operation.operation_name, "operation_number": db_operation.operation_number}
    )

    db_operation = (
        db.query(OperationModel)
        .options(joinedload(OperationModel.user))
        .filter(OperationModel.id == db_operation.id)
        .first()
    )
    pt = db.query(PartTypeModel).filter(PartTypeModel.id == db_operation.part_type_id).first()
    db_operation.part_type_name = pt.type_name if pt else None
    return db_operation


@router.post("/bulk", response_model=List[Operation], status_code=status.HTTP_201_CREATED)
def create_operations_bulk(operations: List[OperationCreate], db: Session = Depends(get_db)):
    """Create many operations in one request (same validations as single create)."""
    if not operations:
        return []

    part_ids = {op.part_id for op in operations if op.part_id}
    if not part_ids or len(part_ids) != 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="bulk create requires exactly one part_id across all operations",
        )
    part_id = next(iter(part_ids))

    # Check if part is scheduled (status is "active")
    schedule_status = db.execute(
        text("SELECT status FROM scheduling.part_schedule_status WHERE part_id = :pid"),
        {"pid": part_id}
    ).fetchone()
    
    if schedule_status and schedule_status[0] and schedule_status[0].lower() == "active":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Part is currently scheduled for production. Operations cannot be added, modified, or deleted while the part is in active production schedule."
        )

    existing_ops = db.query(OperationModel).filter(OperationModel.part_id == part_id).all()
    max_num = 0
    for op in existing_ops:
        try:
            n = int(str(op.operation_number).strip())
        except (TypeError, ValueError):
            continue
        max_num = max(max_num, n)
    next_num = max_num + 10 if max_num > 0 else 10

    requested_numbers: set[str] = set()
    created_ids: List[int] = []

    try:
        for op_in in operations:
            data = op_in.model_dump()
            pt_id = data.get("part_type_id") or 1
            data["part_type_id"] = pt_id

            setup_time_val = data.get("setup_time")
            cycle_time_val = data.get("cycle_time")
            if pt_id != 2:
                if _is_zero_hms(setup_time_val) or _is_zero_hms(cycle_time_val):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="setup_time and cycle_time are mandatory and cannot be 00:00:00 for non Out-Source operations",
                    )

            if not data.get("part_id"):
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="part_id is required for operations")

            if pt_id == 2:
                if not data.get("from_date") or not data.get("to_date"):
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="Outsource operations require from_date and to_date",
                    )

            op_number_raw = data.get("operation_number")
            op_number = op_number_raw.strip() if isinstance(op_number_raw, str) else None
            if op_number:
                if op_number in requested_numbers:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Duplicate operation number '{op_number}' in bulk request",
                    )
                existing = (
                    db.query(OperationModel)
                    .filter(OperationModel.part_id == part_id, OperationModel.operation_number == op_number)
                    .first()
                )
                if existing:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Operation number '{op_number}' already exists for this part",
                    )
                data["operation_number"] = op_number
                requested_numbers.add(op_number)
            else:
                while str(next_num) in requested_numbers:
                    next_num += 10
                data["operation_number"] = str(next_num)
                requested_numbers.add(str(next_num))
                next_num += 10

            db_operation = OperationModel(**data)
            db.add(db_operation)
            db.flush()
            created_ids.append(db_operation.id)

        db.commit()

        created = (
            db.query(OperationModel)
            .options(joinedload(OperationModel.user))
            .filter(OperationModel.id.in_(created_ids))
            .order_by(OperationModel.id.asc())
            .all()
        )
        pt_ids = {o.part_type_id for o in created if o.part_type_id is not None}
        pt_map = {}
        if pt_ids:
            pts = db.query(PartTypeModel).filter(PartTypeModel.id.in_(pt_ids)).all()
            pt_map = {p.id: p.type_name for p in pts}
        for op in created:
            op.part_type_name = pt_map.get(op.part_type_id)
        return created

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create operations: {str(e)}")


@router.get("/", response_model=List[Operation])
def get_operations(user_id: int | None = None, db: Session = Depends(get_db)):
    query = (
        db.query(OperationModel)
        .options(joinedload(OperationModel.user))
        .order_by(OperationModel.id.asc())
    )
    if user_id is not None:
        query = query.filter(OperationModel.user_id == user_id)
    operations = query.all()

    work_center_ids = {op.workcenter_id for op in operations if op.workcenter_id is not None}
    machine_ids = {op.machine_id for op in operations if op.machine_id is not None}
    part_type_ids = {op.part_type_id for op in operations if op.part_type_id is not None}
    vendor_ids = {op.vendor_id for op in operations if op.vendor_id is not None}
    work_center_map = {}
    machine_map = {}
    part_type_map = {}
    vendor_map = {}
    if work_center_ids:
        work_centers = db.query(workcenterModel).filter(workcenterModel.id.in_(work_center_ids)).all()
        work_center_map = {wc.id: wc.work_center_name for wc in work_centers}
    if machine_ids:
        machines = db.query(MachineModel).filter(MachineModel.id.in_(machine_ids)).all()
        machine_map = {m.id: m.make for m in machines}
    if part_type_ids:
        part_types = db.query(PartTypeModel).filter(PartTypeModel.id.in_(part_type_ids)).all()
        part_type_map = {pt.id: pt.type_name for pt in part_types}
    if vendor_ids:
        vendors = db.query(Vendors).filter(Vendors.id.in_(vendor_ids)).all()
        vendor_map = {v.id: v.company_name for v in vendors}

    for op in operations:
        op.work_center_name = work_center_map.get(op.workcenter_id)
        op.machine_name = machine_map.get(op.machine_id)
        op.part_type_name = part_type_map.get(op.part_type_id)
        op.vendor_name = vendor_map.get(op.vendor_id)

    return operations


@router.get("/{operation_id}", response_model=Operation)
def get_operation(operation_id: int, db: Session = Depends(get_db)):
    operation = (
        db.query(OperationModel)
        .options(joinedload(OperationModel.user))
        .filter(OperationModel.id == operation_id)
        .first()
    )
    if not operation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Operation with id {operation_id} not found"
        )
    work_center = None
    if operation.workcenter_id is not None:
        work_center = db.query(workcenterModel).filter(workcenterModel.id == operation.workcenter_id).first()
    machine = None
    if operation.machine_id is not None:
        machine = db.query(MachineModel).filter(MachineModel.id == operation.machine_id).first()
    part_type = None
    if operation.part_type_id is not None:
        part_type = db.query(PartTypeModel).filter(PartTypeModel.id == operation.part_type_id).first()
    
    vendor_name = None
    if operation.vendor_id:
        vendor = db.query(Vendors).filter(Vendors.id == operation.vendor_id).first()
        vendor_name = vendor.company_name if vendor else None
    operation.vendor_name = vendor_name
    
    operation.part_type_name = part_type.type_name if part_type else None
    operation.work_center_name = work_center.work_center_name if work_center else None
    operation.machine_name = machine.make if machine else None
    return operation


@router.get("/part/{part_id}", response_model=List[Operation])
def get_operations_by_part(part_id: int, user_id: int | None = None, db: Session = Depends(get_db)):
    query = (
        db.query(OperationModel)
        .options(
            joinedload(OperationModel.user),
            joinedload(OperationModel.operation_documents),
            joinedload(OperationModel.tools).joinedload(ToolWithPartModel.tool)
        )
        .filter(OperationModel.part_id == part_id)
        .order_by(OperationModel.id.asc())
    )
    if user_id is not None:
        query = query.filter(OperationModel.user_id == user_id)
    operations = query.all()

    work_center_ids = {op.workcenter_id for op in operations if op.workcenter_id is not None}
    machine_ids = {op.machine_id for op in operations if op.machine_id is not None}
    part_type_ids = {op.part_type_id for op in operations if op.part_type_id is not None}
    vendor_ids = {op.vendor_id for op in operations if op.vendor_id is not None}
    work_center_map = {}
    machine_map = {}
    part_type_map = {}
    vendor_map = {}
    if work_center_ids:
        work_centers = db.query(workcenterModel).filter(workcenterModel.id.in_(work_center_ids)).all()
        work_center_map = {wc.id: wc.work_center_name for wc in work_centers}
    if machine_ids:
        machines = db.query(MachineModel).filter(MachineModel.id.in_(machine_ids)).all()
        machine_map = {m.id: m.make for m in machines}
    if part_type_ids:
        part_types = db.query(PartTypeModel).filter(PartTypeModel.id.in_(part_type_ids)).all()
        part_type_map = {pt.id: pt.type_name for pt in part_types}
    if vendor_ids:
        vendors = db.query(Vendors).filter(Vendors.id.in_(vendor_ids)).all()
        vendor_map = {v.id: v.company_name for v in vendors}

    for op in operations:
        op.work_center_name = work_center_map.get(op.workcenter_id)
        op.machine_name = machine_map.get(op.machine_id)
        op.part_type_name = part_type_map.get(op.part_type_id)
        op.vendor_name = vendor_map.get(op.vendor_id)
        if op.tools:
            op.tools.sort(key=lambda x: x.id)

    return operations


@router.put("/{operation_id}", response_model=Operation)
def update_operation(operation_id: int, operation: OperationUpdate, db: Session = Depends(get_db)):
    db_operation = db.query(OperationModel).filter(OperationModel.id == operation_id).first()
    if not db_operation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Operation with id {operation_id} not found"
        )

    # Check if part is scheduled (status is "active")
    if db_operation.part_id:
        schedule_status = db.execute(
            text("SELECT status FROM scheduling.part_schedule_status WHERE part_id = :pid"),
            {"pid": db_operation.part_id}
        ).fetchone()
        
        if schedule_status and schedule_status[0] and schedule_status[0].lower() == "active":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Sorry, this operation cannot be added/modified/deleted because the part is currently scheduled for production. To make changes, please inactivate the part's schedule status first."
            )

    update_data = operation.model_dump(exclude_unset=True)

    # Capture old values before updating
    old_values = {}
    for field in update_data.keys():
        old_values[field] = getattr(db_operation, field, None)

    if "operation_number" in update_data:
        new_op_num_raw = update_data.get("operation_number")
        new_op_num = new_op_num_raw.strip() if isinstance(new_op_num_raw, str) else None
        if not new_op_num:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="operation_number cannot be empty",
            )
        existing = (
            db.query(OperationModel)
            .filter(
                OperationModel.part_id == db_operation.part_id,
                OperationModel.operation_number == new_op_num,
                OperationModel.id != operation_id,
            )
            .first()
        )
        if existing:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Operation number '{new_op_num}' already exists for this part",
            )
        update_data["operation_number"] = new_op_num

    part_type_id = update_data.get("part_type_id") if "part_type_id" in update_data else db_operation.part_type_id
    if part_type_id == 2:
        from_date = update_data.get("from_date") if "from_date" in update_data else db_operation.from_date
        to_date = update_data.get("to_date") if "to_date" in update_data else db_operation.to_date
        if not from_date or not to_date:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Outsource operations require from_date and to_date",
            )

    new_setup = update_data.get("setup_time") if "setup_time" in update_data else db_operation.setup_time
    new_cycle = update_data.get("cycle_time") if "cycle_time" in update_data else db_operation.cycle_time
    if part_type_id != 2:
        if _is_zero_hms(new_setup) or _is_zero_hms(new_cycle):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="setup_time and cycle_time are mandatory and cannot be 00:00:00 for non Out-Source operations",
            )
    for field, value in update_data.items():
        setattr(db_operation, field, value)

    db.commit()
    
    # Log operation update for PC notifications with old and new values
    user_name = None
    user_role = None
    if db_operation.user_id:
        user = db.query(AccessUser).filter(AccessUser.id == db_operation.user_id).first()
        user_name = user.user_name if user else None
        user_role = user.role if user else None
    
    # Capture changes with old and new values
    changes = {}
    for field in update_data.keys():
        old_value = old_values[field]
        new_value = update_data[field]
        
        # Convert time/datetime objects to strings for JSON serialization
        if old_value is not None and hasattr(old_value, 'isoformat'):
            old_value = old_value.isoformat()
        elif hasattr(old_value, '__str__') and not isinstance(old_value, (str, int, float, bool)):
            old_value = str(old_value)
        
        if new_value is not None and hasattr(new_value, 'isoformat'):
            new_value = new_value.isoformat()
        elif hasattr(new_value, '__str__') and not isinstance(new_value, (str, int, float, bool)):
            new_value = str(new_value)
        
        changes[field] = {"old": old_value, "new": new_value}
    
    NotificationService.log_operation_change(
        db=db,
        operation_id=db_operation.id,
        action="updated",
        user_id=db_operation.user_id,
        user_name=user_name,
        user_role=user_role,
        details={
            "operation_name": db_operation.operation_name,
            "operation_number": db_operation.operation_number,
            "changes": changes
        }
    )
    
    db.refresh(db_operation)
    db_operation = (
        db.query(OperationModel)
        .options(joinedload(OperationModel.user))
        .filter(OperationModel.id == operation_id)
        .first()
    )
    pt = db.query(PartTypeModel).filter(PartTypeModel.id == db_operation.part_type_id).first()
    db_operation.part_type_name = pt.type_name if pt else None
    return db_operation


@router.post("/swap", status_code=status.HTTP_200_OK)
def swap_operation_numbers(op1_id: int, op2_id: int, db: Session = Depends(get_db)):
    """Swap operation numbers between two operations"""
    op1 = db.query(OperationModel).filter(OperationModel.id == op1_id).first()
    op2 = db.query(OperationModel).filter(OperationModel.id == op2_id).first()
    
    if not op1:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Operation with id {op1_id} not found"
        )
    if not op2:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Operation with id {op2_id} not found"
        )
    
    if op1.part_id != op2.part_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Operations must belong to the same part to swap numbers"
        )
    
    # Check if part has active schedule status
    _assert_part_not_schedule_active(op1.part_id, db)
    
    # Swap the operation numbers
    temp_number = op1.operation_number
    op1.operation_number = op2.operation_number
    op2.operation_number = temp_number
    
    db.commit()
    db.refresh(op1)
    db.refresh(op2)
    
    return {"message": "Operation numbers swapped successfully", "op1_id": op1_id, "op2_id": op2_id}


@router.delete("/{operation_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_operation(operation_id: int, db: Session = Depends(get_db)):
    """Delete an operation and all its references across all schemas"""
    db_operation = db.query(OperationModel).filter(OperationModel.id == operation_id).first()
    if not db_operation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Operation with id {operation_id} not found"
        )

    # Check if part is scheduled (status is "active")
    if db_operation.part_id:
        schedule_status = db.execute(
            text("SELECT status FROM scheduling.part_schedule_status WHERE part_id = :pid"),
            {"pid": db_operation.part_id}
        ).fetchone()
        
        if schedule_status and schedule_status[0] and schedule_status[0].lower() == "active":
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Sorry, this operation cannot be added/modified/deleted because the part is currently scheduled for production. To make changes, please inactivate the part's schedule status first."
            )

    # Log operation deletion for PC notifications before deletion
    user_name = None
    user_role = None
    if db_operation.user_id:
        user = db.query(AccessUser).filter(AccessUser.id == db_operation.user_id).first()
        user_name = user.user_name if user else None
        user_role = user.role if user else None
    
    NotificationService.log_operation_change(
        db=db,
        operation_id=db_operation.id,
        action="deleted",
        user_id=db_operation.user_id,
        user_name=user_name,
        user_role=user_role,
        details={"operation_name": db_operation.operation_name, "operation_number": db_operation.operation_number}
    )

    try:
        # 1. Delete operation documents from MinIO and database
        documents = db.query(OperationDocumentModel).filter(OperationDocumentModel.operation_id == operation_id).all()
        minio_client = get_minio_client()

        for doc in documents:
            try:
                if doc.document_url:
                    parsed_url = urlparse(doc.document_url)
                    path_parts = parsed_url.path.lstrip('/').split('/', 1)
                    if len(path_parts) >= 2:
                        bucket_name = path_parts[0]
                        object_name = path_parts[1]
                        minio_client.client.remove_object(bucket_name, object_name)
                    elif not parsed_url.netloc and '/' in doc.document_url:
                        path_parts = doc.document_url.lstrip('/').split('/', 1)
                        if len(path_parts) >= 2:
                            bucket_name = path_parts[0]
                            object_name = path_parts[1]
                            minio_client.client.remove_object(bucket_name, object_name)
            except Exception as e:
                print(f"Error deleting file from MinIO for document {doc.id}: {str(e)}")
            db.delete(doc)

        # 2. Delete tools with part references
        tools = db.query(ToolWithPartModel).filter(ToolWithPartModel.operation_id == operation_id).all()
        for tool in tools:
            db.delete(tool)

        # 3. Delete from oms.process_plans
        db.execute(
            text("DELETE FROM oms.process_plans WHERE operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # 4. Delete from oms.order_schedule_status
        db.execute(
            text("DELETE FROM oms.order_schedule_status WHERE operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # 5. Delete from notifications.inspection_notifications
        db.execute(
            text("DELETE FROM notifications.inspection_notifications WHERE operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # 6. Delete from production_monitoring.machine_live_history
        db.execute(
            text("DELETE FROM production_monitoring.machine_live_history WHERE current_operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # 7. Delete from production_monitoring.machine_live_status
        db.execute(
            text("DELETE FROM production_monitoring.machine_live_status WHERE current_operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # 8. Delete from scheduling.production_logs
        db.execute(
            text("DELETE FROM scheduling.production_logs WHERE operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # 9. Delete from scheduling.machine_schedule
        db.execute(
            text("DELETE FROM scheduling.machine_schedule WHERE operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # 10. Delete from scheduling.operation_status
        db.execute(
            text("DELETE FROM scheduling.operation_status WHERE operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # 11. Delete from scheduling.planned_schedule_items
        db.execute(
            text("DELETE FROM scheduling.planned_schedule_items WHERE operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # 12. Delete from scheduling.rescheduling_items
        db.execute(
            text("DELETE FROM scheduling.rescheduling_items WHERE operation_id = :op_id"),
            {"op_id": operation_id}
        )

        # Finally, delete the operation itself
        db.delete(db_operation)
        db.commit()
        return None

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting operation: {str(e)}"
        )


@router.post("/parse-mpp", response_model=List[OperationPreview])
async def parse_mpp_file(file: UploadFile = File(...)):
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Empty file",
        )
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext == ".csv":
        operations = _parse_csv(content)
    elif ext in (".xlsx", ".xls"):
        operations = _parse_xlsx(content)
    elif ext == ".docx":
        operations = _parse_docx(content)
    elif ext == ".pdf":
        operations = _parse_pdf(content)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Use DOCX, CSV, XLSX, or PDF.",
        )
    if not operations:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Could not extract operations from file. Make sure the file contains columns: Op Number, Operation Name, Setup Time, Cycle Time, Work Instructions, Notes.",
        )
    return operations


# ── Operations Template ──────────────────────────────────────────────────────────

TEMPLATE_OBJECT_NAME = "templates/operations_template.docx"


@router.post("/template/upload")
async def upload_operations_template(file: UploadFile = File(...)):
    """
    Upload the operations template file to MinIO.
    This replaces any existing template file.
    """
    if not file.filename or not file.filename.lower().endswith('.docx'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .docx files are allowed for the operations template"
        )
    
    content = await file.read()
    if not content:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Empty file"
        )
    
    minio_client = get_minio_client()
    
    try:
        # Upload to MinIO with fixed object name
        file_stream = io.BytesIO(content)
        url = minio_client.upload_file(
            file_data=file_stream,
            object_name=TEMPLATE_OBJECT_NAME,
            content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        )
        
        return {
            "message": "Operations template uploaded successfully",
            "url": url
        }
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload template: {str(e)}"
        )


@router.get("/template/download")
async def download_operations_template():
    """
    Download the operations template file from MinIO.
    """
    minio_client = get_minio_client()
    
    try:
        # Check if template exists
        if not minio_client.file_exists(TEMPLATE_OBJECT_NAME):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Operations template not found. Please upload the template first."
            )
        
        # Download from MinIO
        content = minio_client.download_file(TEMPLATE_OBJECT_NAME)
        
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": "attachment; filename=Operations_Template.docx"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to download template: {str(e)}"
        )


@router.get("/template/exists")
async def check_template_exists():
    """
    Check if the operations template exists in MinIO.
    """
    minio_client = get_minio_client()
    
    try:
        exists = minio_client.file_exists(TEMPLATE_OBJECT_NAME)
        return {"exists": exists}
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to check template: {str(e)}"
        )
